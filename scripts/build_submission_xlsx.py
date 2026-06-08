"""Build the COM6003 submission workbook: a professional 3-sheet Excel file.

  Sheet 1  "Raw Dataset"      — the cleaned, PII-free analysis dataset
                                (data/processed/oxford_epc_clean.csv, 49,442 dwellings).
  Sheet 2  "EDA"              — exploratory analysis: band distribution, descriptive
                                statistics, efficiency by property type, cleaning log,
                                and embedded charts. All aggregates are live Excel
                                FORMULAS over Sheet 1 (not hardcoded).
  Sheet 3  "Final Prediction" — the champion model's per-certificate predictions
                                (reports/predictions_oxford.csv, 76,400 rows) with a
                                formula-driven performance summary.

Run:  python scripts/build_submission_xlsx.py
Output: Oxford_EPC_Submission.xlsx  (project root)
"""
from __future__ import annotations
import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "Oxford_EPC_Submission.xlsx"
FIG = ROOT / "reports" / "figures"
RATING_ORDER = list("ABCDEFG")

# ---- palette (Oxford Blue brand) -------------------------------------------
INK = "002147"; GOLD = "B0883C"; PAPER = "F3EFE4"; HDR = "1F3A5F"; HAIR = "D9D1BE"
WHITE = "FFFFFF"
thin = Side(style="thin", color=HAIR)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def main() -> int:
    print("Loading data ...")
    raw = pd.read_csv(ROOT / "data" / "processed" / "oxford_epc_clean.csv", low_memory=False)
    pred = pd.read_csv(ROOT / "reports" / "predictions_oxford.csv", low_memory=False)
    rob = json.loads((ROOT / "reports" / "champion_robustness.json").read_text())
    art = json.loads((ROOT / "reports" / "champion_artefact.json").read_text())
    pa = json.loads((ROOT / "reports" / "property_analysis.json").read_text())
    dq = pd.read_csv(ROOT / "reports" / "data_quality_report.csv")
    ov = rob.get("overall", {})

    n_raw, n_raw_cols = raw.shape
    n_pred = len(pred)
    RAW_HDR = 3                       # header row on Sheet 1
    RAW_D0, RAW_D1 = RAW_HDR + 1, RAW_HDR + n_raw          # data rows
    PRED_HDR = 15                     # header row on Sheet 3
    PRED_D0, PRED_D1 = PRED_HDR + 1, PRED_HDR + n_pred
    raw_last = get_column_letter(n_raw_cols)
    pred_last = get_column_letter(pred.shape[1])

    def rawcol(name):                 # Excel column letter of a Sheet-1 field
        return get_column_letter(raw.columns.get_loc(name) + 1)

    def predcol(name):
        return get_column_letter(pred.columns.get_loc(name) + 1)

    print("Writing bulk data sheets ...")
    with pd.ExcelWriter(OUT, engine="openpyxl") as xw:
        raw.to_excel(xw, sheet_name="Raw Dataset", index=False, startrow=RAW_HDR - 1)
        pred.to_excel(xw, sheet_name="Final Prediction", index=False, startrow=PRED_HDR - 1)

    print("Styling + building EDA ...")
    wb = load_workbook(OUT)
    # global default font -> Arial (cheap: applies to all unstyled cells)
    wb._named_styles["Normal"].font = Font(name="Arial", size=10)

    def banner(ws, text, sub, lastcol):
        ws.merge_cells(f"A1:{lastcol}1")
        c = ws["A1"]; c.value = text
        c.font = Font(name="Arial", size=15, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=INK)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 30
        ws.merge_cells(f"A2:{lastcol}2")
        s = ws["A2"]; s.value = sub
        s.font = Font(name="Arial", size=9, italic=True, color="555555")
        s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 16

    def style_header(ws, row, ncols):
        for j in range(1, ncols + 1):
            c = ws.cell(row=row, column=j)
            c.font = Font(name="Arial", size=9, bold=True, color=WHITE)
            c.fill = PatternFill("solid", fgColor=HDR)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER
        ws.row_dimensions[row].height = 26

    # ---------------- Sheet 1: Raw Dataset ----------------------------------
    ws1 = wb["Raw Dataset"]
    banner(ws1, "Oxford Domestic EPC  —  Cleaned Analysis Dataset (LA E07000178)",
           f"{n_raw:,} deduplicated dwellings · {n_raw_cols} fields · personal identifiers "
           f"(address, postcode, UPRN) removed per the data card · "
           f"source: data/processed/oxford_epc_clean.csv", raw_last)
    style_header(ws1, RAW_HDR, n_raw_cols)
    ws1.freeze_panes = f"B{RAW_HDR + 1}"
    ws1.auto_filter.ref = f"A{RAW_HDR}:{raw_last}{RAW_D1}"
    for j in range(1, n_raw_cols + 1):
        ws1.column_dimensions[get_column_letter(j)].width = 16

    # ---------------- Sheet 3: Final Prediction -----------------------------
    ws3 = wb["Final Prediction"]
    banner(ws3, "Final Predictions  —  Champion Model (SAP-stratified gradient boosting)",
           f"{n_pred:,} certificate-level predictions · strict post-2022 temporal hold-out · "
           f"frozen artefact sha256 {art.get('pickle_sha256','')[:12]} ({art.get('frozen_at','')[:10]})",
           pred_last)
    sc, bc, ec = predcol("split"), predcol("band_correct"), predcol("abs_score_error")
    srange = f"${sc}${PRED_D0}:${sc}${PRED_D1}"
    brange = f"${bc}${PRED_D0}:${bc}${PRED_D1}"
    erange = f"${ec}${PRED_D0}:${ec}${PRED_D1}"
    summary = [
        ("Performance summary (hold-out / test split)", "", ""),
        ("Test-set dwellings", f'=COUNTIF({srange},"test")', "0"),
        ("Band accuracy (exact)", f'=COUNTIFS({srange},"test",{brange},TRUE)/COUNTIF({srange},"test")', "0.0%"),
        ("Mean abs. SAP-score error", f'=AVERAGEIFS({erange},{srange},"test")', "0.00"),
        ("Quadratic Weighted Kappa (QWK)", ov.get("qwk"), "0.000"),
        ("Within 15 SAP points", ov.get("within_15"), "0.0%"),
        ("Macro-F1", ov.get("macro_f1"), "0.000"),
        ("MAE (band units)", ov.get("mae_band_units"), "0.000"),
    ]
    ws3["A3"].value = summary[0][0]
    ws3["A3"].font = Font(name="Arial", size=11, bold=True, color=INK)
    for i, (lab, val, fmt) in enumerate(summary[1:], start=4):
        ws3.cell(row=i, column=1, value=lab).font = Font(name="Arial", size=10)
        vc = ws3.cell(row=i, column=2, value=val)
        vc.font = Font(name="Arial", size=10, bold=True, color=INK)
        if fmt:
            vc.number_format = fmt
    ws3.cell(row=12, column=1,
             value="QWK / within-15 / macro-F1 / MAE source: reports/champion_robustness.json"
             ).font = Font(name="Arial", size=8, italic=True, color="777777")
    ws3.cell(row=13, column=1,
             value="Accuracy & mean error above are live formulas over the test rows below."
             ).font = Font(name="Arial", size=8, italic=True, color="777777")
    style_header(ws3, PRED_HDR, pred.shape[1])
    ws3.freeze_panes = f"A{PRED_HDR + 1}"
    ws3.auto_filter.ref = f"A{PRED_HDR}:{pred_last}{PRED_D1}"
    widths3 = {"LMK_KEY": 26, "UPRN": 13, "INSPECTION_DATE": 14, "REPORT_TYPE": 12,
               "actual_band": 11, "actual_sap_score": 14, "predicted_sap_score": 16,
               "predicted_band": 13, "band_correct": 12, "abs_score_error": 13,
               "confidence_proxy": 14, "split": 9}
    for name, w in widths3.items():
        if name in pred.columns:
            ws3.column_dimensions[predcol(name)].width = w

    # ---------------- Sheet 2: EDA (inserted at position 2) -----------------
    ws2 = wb.create_sheet("EDA", index=1)
    banner(ws2, "Exploratory Data Analysis  —  Oxford Domestic EPC",
           "All counts, percentages and statistics below are live Excel formulas computed "
           "directly over the 'Raw Dataset' sheet.", "H")
    rt = rawcol("CURRENT_ENERGY_RATING")
    rt_rng = f"'Raw Dataset'!${rt}${RAW_D0}:${rt}${RAW_D1}"

    def sec(row, text):
        ws2.merge_cells(f"A{row}:E{row}")
        c = ws2[f"A{row}"]; c.value = text
        c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=GOLD)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws2.row_dimensions[row].height = 20

    def hrow(row, labels, start=1):
        for k, lab in enumerate(labels):
            c = ws2.cell(row=row, column=start + k, value=lab)
            c.font = Font(name="Arial", size=9, bold=True, color=WHITE)
            c.fill = PatternFill("solid", fgColor=HDR)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER

    r = 4
    # 1) overview
    sec(r, "1 · Dataset overview"); r += 1
    overview = [("Dwellings (deduplicated)", f'=COUNTA({rt_rng})', "0"),
                ("Fields", n_raw_cols, "0"),
                ("Local authority", "Oxford (E07000178)", ""),
                ("Inspection date range", "2007-08-09  to  2026-02-28", ""),
                ("Target variable", "CURRENT_ENERGY_RATING (EPC band A–G)", ""),
                ("Source file", "data/processed/oxford_epc_clean.csv", "")]
    for lab, val, fmt in overview:
        ws2.cell(row=r, column=1, value=lab).font = Font(name="Arial", size=10)
        vc = ws2.cell(row=r, column=2, value=val); vc.font = Font(name="Arial", size=10, bold=True)
        if fmt:
            vc.number_format = fmt
        r += 1
    r += 1

    # 2) band distribution (COUNTIF over Sheet 1)
    sec(r, "2 · EPC band distribution"); r += 1
    hrow(r, ["EPC band", "Dwellings", "Share"]); r += 1
    band_top = r
    for b in RATING_ORDER:
        ws2.cell(row=r, column=1, value=b).font = Font(name="Arial", size=10, bold=True)
        ws2.cell(row=r, column=2, value=f'=COUNTIF({rt_rng},"{b}")').number_format = "#,##0"
        ws2.cell(row=r, column=3, value=f"=B{r}/$B${band_top + 7}").number_format = "0.0%"
        for col in (1, 2, 3):
            ws2.cell(row=r, column=col).border = BORDER
        r += 1
    ws2.cell(row=r, column=1, value="Total").font = Font(name="Arial", size=10, bold=True)
    ws2.cell(row=r, column=2, value=f"=SUM(B{band_top}:B{r-1})").font = Font(name="Arial", size=10, bold=True)
    ws2.cell(row=r, column=2).number_format = "#,##0"
    ws2.cell(row=r, column=3, value=f"=B{r}/$B${r}").number_format = "0.0%"
    r += 2

    # 3) descriptive statistics (formulas over Sheet 1 numeric columns)
    sec(r, "3 · Descriptive statistics (key numeric fields)"); r += 1
    hrow(r, ["Field", "Count", "Mean", "Std dev", "Min", "25%", "Median", "75%", "Max"]); r += 1
    stat_fields = [("CURRENT_ENERGY_EFFICIENCY", "SAP score (current)"),
                   ("TOTAL_FLOOR_AREA", "Total floor area (m²)"),
                   ("NUMBER_HABITABLE_ROOMS", "Habitable rooms"),
                   ("CO2_EMISSIONS_CURRENT", "CO₂ emissions (t/yr)"),
                   ("ENERGY_CONSUMPTION_CURRENT", "Energy use (kWh/m²/yr)"),
                   ("ENVIRONMENTAL_IMPACT_CURRENT", "Environmental impact")]
    for field, nice in stat_fields:
        if field not in raw.columns:
            continue
        cl = rawcol(field)
        rng = f"'Raw Dataset'!${cl}${RAW_D0}:${cl}${RAW_D1}"
        ws2.cell(row=r, column=1, value=nice).font = Font(name="Arial", size=10)
        formulas = [f"=COUNT({rng})", f"=AVERAGE({rng})", f"=STDEV({rng})", f"=MIN({rng})",
                    f"=QUARTILE({rng},1)", f"=MEDIAN({rng})", f"=QUARTILE({rng},3)", f"=MAX({rng})"]
        for k, f in enumerate(formulas, start=2):
            c = ws2.cell(row=r, column=k, value=f)
            c.number_format = "#,##0" if k == 2 else "#,##0.0"
            c.border = BORDER
        ws2.cell(row=r, column=1).border = BORDER
        r += 1
    r += 1

    # 4) efficiency by property type (AVERAGEIF / COUNTIF over Sheet 1)
    sec(r, "4 · Mean SAP score by property type"); r += 1
    hrow(r, ["Property type", "Dwellings", "Mean SAP", "Modal band"]); r += 1
    pt = rawcol("PROPERTY_TYPE"); sap = rawcol("CURRENT_ENERGY_EFFICIENCY")
    pt_rng = f"'Raw Dataset'!${pt}${RAW_D0}:${pt}${RAW_D1}"
    sap_rng = f"'Raw Dataset'!${sap}${RAW_D0}:${sap}${RAW_D1}"
    for d in pa["property_type_efficiency"]["by_type"]:
        t = d["property_type"]
        ws2.cell(row=r, column=1, value=t).font = Font(name="Arial", size=10)
        ws2.cell(row=r, column=2, value=f'=COUNTIF({pt_rng},"{t}")').number_format = "#,##0"
        ws2.cell(row=r, column=3, value=f'=AVERAGEIF({pt_rng},"{t}",{sap_rng})').number_format = "#,##0.0"
        ws2.cell(row=r, column=4, value=d["modal_band"]).font = Font(name="Arial", size=10, bold=True)
        for col in (1, 2, 3, 4):
            ws2.cell(row=r, column=col).border = BORDER
        r += 1
    ws2.cell(row=r, column=4, value="modal band: source reports/property_analysis.json").font = \
        Font(name="Arial", size=8, italic=True, color="777777")
    r += 2

    # 5) data cleaning / quality log
    sec(r, "5 · Data cleaning & quality log"); r += 1
    hrow(r, ["Validation rule", "Rows affected", "Action taken"]); r += 1
    for _, row_ in dq.iterrows():
        ws2.cell(row=r, column=1, value=str(row_["rule"])).font = Font(name="Arial", size=10)
        ws2.cell(row=r, column=2, value=int(row_["n_rows_affected"])).number_format = "#,##0"
        ws2.cell(row=r, column=3, value=str(row_["action"])).font = Font(name="Arial", size=10)
        for col in (1, 2, 3):
            ws2.cell(row=r, column=col).border = BORDER
        r += 1

    # column widths + embedded charts
    for col, w in {"A": 30, "B": 13, "C": 12, "D": 12, "E": 11, "F": 11, "G": 11, "H": 11, "I": 11}.items():
        ws2.column_dimensions[col].width = w
    img_anchor = 4
    for fig_name, anchor in [("fig_rating_distribution", "K"), ("fig_sap_histogram", "K"),
                             ("fig_property_efficiency", "K")]:
        p = FIG / f"{fig_name}.png"
        if p.exists():
            im = XLImage(str(p))
            scale = 470 / im.width
            im.width = int(im.width * scale); im.height = int(im.height * scale)
            ws2.add_image(im, f"{anchor}{img_anchor}")
            img_anchor += 24
    ws2.sheet_view.showGridLines = False
    ws1.sheet_view.showGridLines = False
    ws3.sheet_view.showGridLines = False

    # Force Excel to recalculate every formula on open (openpyxl writes formula
    # strings with no cached value; without this they can show 0 until F9).
    wb.calculation.fullCalcOnLoad = True

    wb.save(OUT)
    print(f"Wrote {OUT.relative_to(ROOT)}  ({OUT.stat().st_size/1e6:.1f} MB)")
    print("Sheets:", wb.sheetnames)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
